import os
from openpyxl import load_workbook
from file_paths.file_paths import path_to_res


# TODO: оформить в тест, добавить проверки и использовать универсальный путь

def test_xlsx_file():
    workbook = load_workbook(os.path.join(path_to_res, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active

    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
